import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

CRM = r'C:\Users\Noraxon\Downloads\CRM 2.0.xlsx'
wb = load_workbook(CRM, data_only=False)
sheet = wb["2' - Clientes"]
cell = sheet.cell(row=7, column=5)
v = cell.value
if isinstance(v, ArrayFormula):
    print(v.text)
